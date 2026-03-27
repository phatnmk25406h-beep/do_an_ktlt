import openpyxl
from datetime import datetime
import os
from src.config.settings import BASE_DIR

EXCEL_PATH = os.path.join(BASE_DIR, "data", "users_data.xlsx")

def get_user_dashboard_data(username):
    """Lấy dữ liệu 7 ngày và thông tin cơ bản để show lên Account Page"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    # Cấu trúc cột mặc định (dựa theo thứ tự bạn tự setup)
    # Ví dụ: Username ở cột 1, Mon ở cột 8, Sun ở cột 14
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == username:
            # Lấy 7 cột ngày (Giả sử từ cột H(8) đến N(14))
            # Nếu giá trị None thì mặc định là 0
            minutes = [float(cell.value) if cell.value else 0.0 for cell in row[7:14]]
            
            # Trả về data
            return {
                "username": username,
                "level": row[5].value or 1,
                "minutes": minutes
            }
    return None

def add_learning_time(username, minutes_spent):
    """Hàm này sẽ được gọi khi người dùng thoát khỏi 1 bài học"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    now = datetime.now()
    current_weekday = now.weekday() # 0: Mon, 1: Tue, ..., 6: Sun
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == username:
            # --- KIỂM TRA RESET TUẦN MỚI ---
            last_date_str = row[6].value # Giả sử Cột G(7) lưu ngày
            if last_date_str:
                last_date = datetime.strptime(str(last_date_str), "%Y-%m-%d")
                # Nếu khác tuần (dựa vào hàm isocalendar), reset 7 cột ngày về 0
                if now.isocalendar()[1] != last_date.isocalendar()[1]:
                    for i in range(7, 14):
                        row[i].value = 0.0
            
            # --- CẬP NHẬT THỜI GIAN HÔM NAY ---
            row[6].value = now.strftime("%Y-%m-%d") # Cập nhật Last_Update_Date
            
            # Cột tương ứng với hôm nay: (Current Weekday + Cột bắt đầu là 8)
            # Index trong mảng row là: 7 + current_weekday
            col_index = 7 + current_weekday 
            current_time = float(row[col_index].value) if row[col_index].value else 0.0
            row[col_index].value = current_time + minutes_spent
            
            wb.save(EXCEL_PATH)
            return True
    return False